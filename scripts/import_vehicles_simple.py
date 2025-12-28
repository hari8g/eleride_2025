#!/usr/bin/env python3
"""
Simple vehicle import script using openpyxl directly.
Run this from the project root with: python3 scripts/import_vehicles_simple.py
"""

import json
import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root / "services" / "platform-api"))

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Error: openpyxl not installed")
    print("Install it with: pip install openpyxl")
    sys.exit(1)

try:
    import requests
except ImportError:
    print("❌ Error: requests not installed")
    print("Install it with: pip install requests")
    sys.exit(1)


def normalize_col_name(col: str) -> str:
    """Normalize column name."""
    return str(col).strip().lower().replace(" ", "_").replace("-", "_")


def get_auth_token(api_url: str, phone: str, operator_slug: str) -> str:
    """Authenticate and get token."""
    print(f"🔐 Requesting OTP for {phone}...")
    otp_resp = requests.post(
        f"{api_url}/operator/auth/otp/request",
        json={"phone": phone, "mode": "login", "operator_slug": operator_slug},
        timeout=30,
    )
    if otp_resp.status_code != 200:
        raise Exception(f"OTP request failed: {otp_resp.text}")
    
    otp_data = otp_resp.json()
    request_id = otp_data["request_id"]
    dev_otp = otp_data.get("dev_otp")
    
    if not dev_otp:
        otp = input("Enter OTP manually: ").strip()
    else:
        otp = dev_otp
        print(f"✅ Dev OTP: {otp}")
    
    print("🔐 Verifying OTP...")
    verify_resp = requests.post(
        f"{api_url}/operator/auth/otp/verify",
        json={"request_id": request_id, "otp": otp},
        timeout=30,
    )
    if verify_resp.status_code != 200:
        raise Exception(f"OTP verification failed: {verify_resp.text}")
    
    session = verify_resp.json()
    print("✅ Authenticated successfully\n")
    return session["access_token"]


def create_vehicle(api_url: str, token: str, vehicle_data: dict) -> dict:
    """Create vehicle via API."""
    response = requests.post(
        f"{api_url}/operator/vehicles",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json=vehicle_data,
        timeout=30,
    )
    if response.status_code not in (200, 201):
        raise Exception(f"{response.status_code}: {response.text}")
    return response.json()


def main():
    excel_path = project_root / "data" / "Vehicle Inventory Data-271225xlsx.xlsx"
    if not excel_path.exists():
        print(f"❌ Error: Excel file not found: {excel_path}")
        sys.exit(1)
    
    print(f"📖 Reading Excel file: {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # Get headers
    headers = [normalize_col_name(cell.value or "") for cell in ws[1]]
    print(f"✅ Found {len(headers)} columns")
    print(f"   Columns: {', '.join(headers[:10])}...\n")
    
    # Parse rows
    vehicles = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                value = cell.value
                if value is not None:
                    row_data[header] = str(value).strip()
        
        # Skip empty rows
        if not row_data:
            continue
        
        # Extract key fields
        vin = row_data.get("vin", "").strip()
        reg_number = row_data.get("reg_number", "").strip() or row_data.get("registration_number", "").strip()
        brand = row_data.get("brand", "").strip() or row_data.get("make", "").strip()
        model = row_data.get("model", "").strip()
        variant = row_data.get("variant", "").strip() or row_data.get("trim", "").strip()
        color = row_data.get("color", "").strip()
        reg_date = row_data.get("reg_date", "").strip() or row_data.get("registration_date", "").strip()
        chassis = row_data.get("chassis_number", "").strip() or row_data.get("chassis", "").strip()
        mfg_year = row_data.get("mfg_year", "").strip() or row_data.get("year", "").strip()
        owner_name = row_data.get("rc_owner_name", "").strip() or row_data.get("owner_name", "").strip()
        battery_id = row_data.get("battery_id", "").strip()
        battery_number = row_data.get("battery_number", "").strip()
        location = row_data.get("location", "").strip()
        city = row_data.get("city", "").strip()
        
        # Use VIN as primary identifier
        if not vin and not reg_number:
            print(f"⚠️  Row {row_idx}: Skipping - no VIN or registration number")
            continue
        
        identifier = vin if vin else reg_number
        
        # Build metadata
        meta = {
            "vin": vin or reg_number,
            "registration_number": reg_number or vin,
            "brand": brand,
            "model": model,
            "variant": variant,
            "color": color,
            "reg_date": reg_date,
            "chassis_number": chassis,
            "mfg_year": mfg_year,
            "rc_owner_name": owner_name,
            "battery_id": battery_id,
            "battery_number": battery_number,
            "location": location,
            "city": city,
        }
        # Remove empty values
        meta = {k: v for k, v in meta.items() if v}
        
        vehicles.append({
            "registration_number": reg_number if reg_number else vin,  # Use registration number, fallback to VIN
            "vin": vin,  # Send VIN explicitly
            "model": f"{brand} {model}".strip() if (brand or model) else None,
            "meta": json.dumps(meta),
        })
    
    print(f"✅ Parsed {len(vehicles)} vehicles\n")
    
    # Show sample
    if vehicles:
        print("Sample vehicle:")
        sample = vehicles[0]
        meta_sample = json.loads(sample["meta"])
        print(f"  VIN: {meta_sample.get('vin', 'N/A')}")
        print(f"  Reg: {meta_sample.get('registration_number', 'N/A')}")
        print(f"  Brand: {meta_sample.get('brand', 'N/A')}")
        print(f"  Model: {meta_sample.get('model', 'N/A')}")
        print()
    
    # Auto-confirm if running non-interactively
    import os
    if os.isatty(0):
        confirm = input(f"Import {len(vehicles)} vehicles? (yes/no): ").strip().lower()
        if confirm != "yes":
            print("❌ Import cancelled")
            return
    else:
        print(f"📦 Auto-importing {len(vehicles)} vehicles (non-interactive mode)...")
    
    # Authenticate
    # Use AWS API if API_URL env var is set, otherwise localhost
    import os
    api_url = os.getenv("API_URL", "http://localhost:18080")
    phone = "+919999000401"
    operator_slug = "eleride-fleet"
    
    try:
        token = get_auth_token(api_url, phone, operator_slug)
    except Exception as e:
        print(f"❌ Authentication failed: {e}")
        sys.exit(1)
    
    # Import vehicles
    created = 0
    failed = 0
    
    print(f"🚀 Importing {len(vehicles)} vehicles...\n")
    
    for i, vehicle_data in enumerate(vehicles, 1):
        try:
            result = create_vehicle(api_url, token, vehicle_data)
            created += 1
            meta = json.loads(vehicle_data["meta"])
            vin = meta.get("vin", vehicle_data["registration_number"])
            print(f"✅ [{i}/{len(vehicles)}] {vin}")
        except Exception as e:
            failed += 1
            vin = json.loads(vehicle_data["meta"]).get("vin", vehicle_data["registration_number"])
            error_msg = str(e)
            if "already exists" in error_msg.lower():
                print(f"⚠️  [{i}/{len(vehicles)}] {vin} - Already exists (skipped)")
                created += 1
                failed -= 1
            else:
                print(f"❌ [{i}/{len(vehicles)}] {vin} - {error_msg}")
    
    print(f"\n📊 Import Summary:")
    print(f"   ✅ Created/Updated: {created}")
    print(f"   ❌ Failed: {failed}")
    print(f"   📦 Total: {len(vehicles)}")
    print(f"\n✅ Done! Check the fleet portal at http://localhost:5177")


if __name__ == "__main__":
    main()

